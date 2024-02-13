import camelot  
import re
import unicodedata
from PyPDF2 import PdfReader
from .models import Product, Category
import openpyxl

def normalize_string(s):
    return unicodedata.normalize('NFKC', s)

def get_point_of_sale(file_path):
    with open(file_path, 'rb') as f:
        pdf = PdfReader(f)
        text = pdf.pages[0].extract_text()
        point_of_sale = text.split("\n")[1]
        return point_of_sale
    
def get_categories_with_products():
    categories_with_products = {}
    categories = Category.objects.all()  # Get all Category instances
    for category in categories:
        # Directly use 'category' instance to filter Products
        products = Product.objects.filter(category=category).values_list('name', flat=True)
        categories_with_products[category.name] = list(products)
    return categories_with_products


def get_product_table(pdf_file):
    table = camelot.read_pdf(pdf_file, pages='3', flavor='stream')
    df = table[0].df
    df = df.dropna()
    df.columns = df.iloc[0]
    df = df[1:] 
    
    products = []

    print(df.columns)
    for _, row in df.iterrows():
        product_name = row["Produkte"]
        net_price = row["Nettopreis"].replace(",", ".").replace("EUR", "").strip()
        gross_price = row["Bruttopreis"].replace(",", ".").replace("EUR", "").strip()
        quantity = row["Menge"]

        if product_name != "":
            if net_price == "" and quantity == "":
                products[-1]["Produkte"] = products[-1]["Produkte"][:-1] + "€ " + product_name
                continue
            elif net_price == "" and quantity != "":
                continue
        else:
            continue

        if len(re.findall("XXL", product_name)) == 2:
            product_name = product_name.replace(" XXL ", " ")

        net_price = float(net_price)
        gross_price = float(gross_price)
        
        products.append({"Produkte": normalize_string(product_name), "Nettopreis": net_price, "Bruttopreis": gross_price, "Menge": int(quantity)})

    return products

def categorize(products):
    categories = get_categories_with_products() 
    categorized_products = {}
    for category in categories:
        categorized_products[category] = []

    for product in products:
        if "einkauf" in product["Produkte"].lower():
            product["Produkte"] = re.sub(r"ab (\d+). einkauf", r"ab \1€ einkauf", product["Produkte"], flags=re.IGNORECASE)
    
    for product in products:
        for category in categories:
            for product_name in categories[category]:
                if re.match(product_name, product["Produkte"], re.IGNORECASE):
                    if "angebot" not in product["Produkte"].lower():
                        product["Produkte"] = product_name
                    categorized_products[category].append(product)
                    break

            if not re.match(product_name, product["Produkte"], re.IGNORECASE):
                categorized_products["Other"]

    return categorized_products

def write_point_of_sale_to_excel(wb, point_of_sale, categories_with_products):
    current_sheet = wb.create_sheet(point_of_sale)
    heading = current_sheet.cell(row=1, column=1, value=f"Einnahmen {point_of_sale}")
    heading.font = openpyxl.styles.Font(size=16, bold=True, name="Calibri")

    h1 = current_sheet.cell(row=3, column=1, value="Produkte")
    h1.font = openpyxl.styles.Font(size=12, name="Calibri", italic=True)
    h1.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
    
    h2 = current_sheet.cell(row=3, column=2, value="Nettopreis")
    h2.font = openpyxl.styles.Font(size=12, name="Calibri", italic=True)
    h2.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
    
    h3 = current_sheet.cell(row=3, column=3, value="Bruttopreis")
    h3.font = openpyxl.styles.Font(size=12, name="Calibri", italic=True)
    h3.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
    
    h4 = current_sheet.cell(row=3, column=4, value="Menge")
    h4.font = openpyxl.styles.Font(size=12, name="Calibri", italic=True)
    h4.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))

    # increase the width of the columns
    current_sheet.column_dimensions["A"].width = 50
    current_sheet.column_dimensions["B"].width = 15
    current_sheet.column_dimensions["C"].width = 15
    current_sheet.column_dimensions["D"].width = 15

    row = 4
    for category, products in categories_with_products.items():
        category_heading = current_sheet.cell(row=row, column=1, value=category)
        category_heading.font = openpyxl.styles.Font(size=11, name="Calibri", bold=True)
        category_heading.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
        category_heading.fill = openpyxl.styles.PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
        last_col = current_sheet.max_column
        current_sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)

        row += 1
        for product in sorted(products, key=lambda x: x["Produkte"]):
            c1 = current_sheet.cell(row=row, column=1, value=product["Produkte"])
            c1.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
            c2 = current_sheet.cell(row=row, column=2, value=product["Nettopreis"])
            c2.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
            c2.number_format = "#,##0.00 €"
            c3 = current_sheet.cell(row=row, column=3, value=product["Bruttopreis"])
            c3.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
            c3.number_format = "#,##0.00 €"
            c4 = current_sheet.cell(row=row, column=4, value=product["Menge"])
            c4.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
            row += 1

def save_data_to_excel(file_list, save_file_path, categories_with_products):
    wb = openpyxl.Workbook()
    main_sheet = wb.active
    main_sheet.title = "Verkauf pro Stadt"
    current_sheet = wb["Verkauf pro Stadt"]

    main_heading = current_sheet.cell(row=1, column=1, value="Auswertung Verkauf")
    main_heading.font = openpyxl.styles.Font(size=16, bold=True, name="Calibri")

    product_heading = current_sheet.cell(row=3, column=1, value="Produkte")
    product_heading.alignment = openpyxl.styles.Alignment(vertical="center")
    product_heading.font = openpyxl.styles.Font(size=12, italic=True, name="Calibri")
    product_heading.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
    
    net_price_heading = current_sheet.cell(row=3, column=2, value="Nettopreis")
    net_price_heading.font = openpyxl.styles.Font(size=12, italic=True, name="Calibri", bold=True)
    net_price_heading.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
    net_price_heading.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    gross_price_heading = current_sheet.cell(row=3, column=3, value="Bruttopreis")
    gross_price_heading.font = openpyxl.styles.Font(size=12, italic=True, name="Calibri", bold=True)
    gross_price_heading.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
    gross_price_heading.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    total_heading = current_sheet.cell(row=3, column=len(file_list)+4, value="Gesamt")
    total_heading.font = openpyxl.styles.Font(size=12, italic=True, name="Calibri", bold=True)
    total_heading.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
    total_heading.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")


    point_of_sale_data = {}
    product_prices = {}
    for file_path in file_list:
        point_of_sale = get_point_of_sale(file_path)
        products = get_product_table(file_path)
        categorized_products = categorize(products)
        point_of_sale_data[point_of_sale] = categorized_products
        for _ , products in categorized_products.items():
            for product in products:
                if product["Produkte"] not in product_prices:
                    product_prices[product["Produkte"]] = [product["Nettopreis"], product["Bruttopreis"]]
                else:
                    if product["Nettopreis"] < product_prices[product["Produkte"]][0]:
                        product_prices[product["Produkte"] + " Angebot"] = [product["Nettopreis"], product["Bruttopreis"]]
                    elif product["Nettopreis"] > product_prices[product["Produkte"]][0]:
                        product_prices[product["Produkte"] + " Angebot"] = [product_prices[product["Produkte"]][0], product_prices[product["Produkte"]][1]]
                        product_prices[product["Produkte"]] = [product["Nettopreis"], product["Bruttopreis"]]
    
    
    product_row_indices = {}
    row = 4
    for category in categories_with_products:
        category_heading = main_sheet.cell(row=row, column=1, value=category)
        category_heading.font = openpyxl.styles.Font(size=14, name="Calibri", bold=True)
        category_heading.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
        category_heading.fill = openpyxl.styles.PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid") 
        row += 1
        products = categories_with_products[category]
        for product in sorted(products):
            if product_prices.get(product):
                product_row_indices[product] = row

                c = main_sheet.cell(row=row, column=1, value=product)
                c.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
                
                c = main_sheet.cell(row=row, column=2, value=product_prices[product][0])
                c.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
                c.number_format = "#,##0.00 €"

                c = main_sheet.cell(row=row, column=3, value=product_prices[product][1])
                c.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
                c.number_format = "#,##0.00 €"
                row += 1

            if product_prices.get(product + " Angebot"):
                product_row_indices[product + " Angebot"] = row

                c = main_sheet.cell(row=row, column=1, value=f"{product} Angebot")
                c.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
                
                c = main_sheet.cell(row=row, column=2, value=product_prices[product + " Angebot"][0])
                c.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
                c.number_format = "#,##0.00 €"

                c = main_sheet.cell(row=row, column=3, value=product_prices[product + " Angebot"][1])
                c.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))     
                c.number_format = "#,##0.00 €"
                row += 1

    col_index = 4
    for point_of_sale in point_of_sale_data:
        categorized_products = point_of_sale_data[point_of_sale]
        write_point_of_sale_to_excel(wb, point_of_sale, categorized_products)
        current_sheet = wb["Verkauf pro Stadt"]

        point_of_sale_heading = current_sheet.cell(row=3, column=col_index, value=point_of_sale)
        point_of_sale_heading.font = openpyxl.styles.Font(size=12, italic=True, name="Calibri")
        point_of_sale_heading.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"))
        point_of_sale_heading.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        for category, products in categorized_products.items():
            for product in products:
                if product["Nettopreis"] < product_prices[product["Produkte"]][0]:
                    row_index = product_row_indices[product["Produkte"] + " Angebot"]
                else:
                    row_index = product_row_indices[product["Produkte"]]

                c = current_sheet.cell(row=row_index, column=col_index, value=product["Menge"])
                c.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"))

        col_index += 1
        
    current_sheet = wb["Verkauf pro Stadt"]
    for row in range(4, current_sheet.max_row+1):
        if current_sheet.cell(row=row, column=1).value in categories_with_products:
            continue
        for col in range(2, current_sheet.max_column):
            if current_sheet.cell(row=row, column=col).value == None:
                current_sheet.cell(row=row, column=col).value = 0
                current_sheet.cell(row=row, column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"))

    for row in range(4, current_sheet.max_row+1):
        if current_sheet.cell(row=row, column=1).value in categories_with_products:
            current_sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=current_sheet.max_column)
            continue
        sum_of_row = 0
        for col in range(4, current_sheet.max_column):
            sum_of_row += current_sheet.cell(row=row, column=col).value
        current_sheet.cell(row=row, column=current_sheet.max_column, value=sum_of_row)
        current_sheet.cell(row=row, column=current_sheet.max_column).font = openpyxl.styles.Font(bold=True)
        current_sheet.cell(row=row, column=current_sheet.max_column).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style="thin"), top=openpyxl.styles.Side(border_style="thin"), right=openpyxl.styles.Side(border_style="thin"), left=openpyxl.styles.Side(border_style="thin"))
    
    # increase the width of the columns
    current_sheet.column_dimensions["A"].width = 50
    for col in range(2, current_sheet.max_column+1):
        if col == 7:
            current_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        else:
            current_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    current_sheet.row_dimensions[3].height = 30

    wb.save(save_file_path)